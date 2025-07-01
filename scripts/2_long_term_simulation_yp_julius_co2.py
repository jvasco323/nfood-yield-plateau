# ----------------------------------------------------------------------------------------------------------------------
# Load libraries
import pcse
from pcse.fileinput import YAMLCropDataProvider
import os
import pandas as pd
import numpy as np
from pcse.base import ParameterProvider
from pcse.fileinput import ExcelWeatherDataProvider
from pcse.models import WOFOST8_ML_PP, WOFOST8_ML_NLP, WOFOST8_ML_WLP, WOFOST8_ML_NWLP
import yaml
from openpyxl import load_workbook

# ----------------------------------------------------------------------------------------------------------------------
# Pre-defined settings
cultivar = "Julius"
location = "Wageningen"
treatment = "N3"
year = 2015
nlim = False
wlim = False

# ----------------------------------------------------------------------------------------------------------------------
# Get directories of input and output files
work_dir = os.getcwd()
input_dir = os.path.join(work_dir, "4-hberghuijs-model--final/input")
output_dir = os.path.join(work_dir, "4-hberghuijs-model--final/output")

# ----------------------------------------------------------------------------------------------------------------------
# Set crop parameters and input data
cropdata = YAMLCropDataProvider(fpath=os.path.join(input_dir, "crop"), force_reload=True)
cropdata.set_active_crop('winterwheat', f"{cultivar}")
sitedata = yaml.safe_load(open(os.path.join(input_dir, "site", f"{location}_{year}_site.yaml")))
soildata_sand = yaml.safe_load(open(os.path.join(input_dir, "soil", f"wageningen_north_soil.yaml")))
soildata_clay = yaml.safe_load(open(os.path.join(input_dir, "soil", f"lelystad_soil.yaml")))

# ----------------------------------------------------------------------------------------------------------------------
# 0. Choose soil type
for soil in ['clay']:
    print(soil)
    if (soil == 'sand'):
        soildata = soildata_sand
        parameters = ParameterProvider(cropdata=cropdata, soildata=soildata, sitedata=sitedata)
    else:
        soildata = soildata_clay
        parameters = ParameterProvider(cropdata=cropdata, soildata=soildata, sitedata=sitedata)

# ----------------------------------------------------------------------------------------------------------------------
# 1. Loop per weather station
    weather_stations = ['DE BILT', 'EELDE', 'VLISSINGEN']
    for station in weather_stations:
        print(station)
        # 2. Select weather station
        weatherdata = ExcelWeatherDataProvider(os.path.join(input_dir, "weather", f"WOFOST_Weather_KNMI_{station}.xlsx"))
        # 3. Get co2 data
        for CO2 in [360, 400]:
            print(CO2)
            Output = pd.DataFrame()
            Summary = pd.DataFrame()
            for yr in np.arange(1972, 2016, 1):
                parameters.set_override("CO2", CO2)
                # 4. Load agromanager
                agrod = yaml.safe_load(open(os.path.join(input_dir, "agro-julius", f"Agrom_WW_Monocrop_{yr}.txt")))
                agrod = agrod['AgroManagement']
                # 5. Choose model configuration
                if (nlim == True and wlim == True):
                    wofost = WOFOST8_ML_NWLP(parameters, weatherdata, agrod)
                elif (nlim == False and wlim == True):
                    wofost = WOFOST8_ML_WLP(parameters, weatherdata, agrod)
                elif (nlim == True and wlim == False):
                    wofost = WOFOST8_ML_NLP(parameters, weatherdata, agrod)
                else:
                    wofost = WOFOST8_ML_PP(parameters, weatherdata, agrod)
                # 6. Run model
                wofost.run_till_terminate()
                # 7. Save model outputs
                Output_PP = pd.DataFrame(wofost.get_output())
                Summary_PP = pd.DataFrame(wofost.get_summary_output())
                # 8. Append years
                Output = pd.concat([Output, Output_PP], ignore_index=True)
                Summary = pd.concat([Summary, Summary_PP], ignore_index=True)
            # 9. Save daily output
            Output['dt_day'] = pd.to_datetime(Output.day)
            weather_df = pd.DataFrame(weatherdata.export())
            weather_df['dt_day'] = pd.to_datetime(weather_df.DAY)
            df_wofost_meteo = pd.merge(Output, weather_df, on='dt_day')
            path = os.path.join(output_dir, f"co2-check/yp-long-term-daily-{station}-{CO2}.xlsx")
            book = load_workbook(path)
            with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_wofost_meteo.to_excel(writer, index=False, sheet_name='Sheet1')
            # 10. Save summary output
            Summary['DOM'] = pd.to_datetime(Summary['DOM'])
            Summary['year'] = Summary['DOM'].dt.year
            path = os.path.join(output_dir, f"co2-check/yp-long-term-summary-{station}-{CO2}.xlsx")
            book = load_workbook(path)
            with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                Summary.to_excel(writer, index=False, sheet_name='Sheet1')
