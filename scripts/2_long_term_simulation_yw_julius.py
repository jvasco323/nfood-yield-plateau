# ----------------------------------------------------------------------------------------------------------------------
# Load libraries
import pcse
from pcse.fileinput import YAMLCropDataProvider
import os
import pandas as pd
from pcse.base import ParameterProvider
from pcse.fileinput import ExcelWeatherDataProvider
from pcse.models import WOFOST8_ML_PP, WOFOST8_ML_WLP, WOFOST8_ML_NLP, WOFOST8_ML_NWLP
import yaml
from openpyxl import load_workbook

# ----------------------------------------------------------------------------------------------------------------------
# Pre-defined settings
cultivar = "Julius"
location = "Wageningen"
treatment = "N3"
year = 2015
nlim = False
wlim = True

# ----------------------------------------------------------------------------------------------------------------------
# Get directories of input and output files
work_dir = os.getcwd()
input_dir = os.path.join(work_dir, "4-hberghuijs-model--final/input")
output_dir = os.path.join(work_dir, "4-hberghuijs-model--final/output")

# ----------------------------------------------------------------------------------------------------------------------
# Set crop parameters and input data
cropdata = YAMLCropDataProvider(fpath=os.path.join(input_dir, "crop"), force_reload=True)
cropdata.set_active_crop('winterwheat', f"{cultivar}")
sitedata = yaml.safe_load(open(os.path.join(input_dir, "site", f"{location}_{year}_site.yaml")))
soildata_sand = yaml.safe_load(open(os.path.join(input_dir, "soil", f"wageningen_north_soil.yaml")))
soildata_clay = yaml.safe_load(open(os.path.join(input_dir, "soil", f"lelystad_soil.yaml")))

# ----------------------------------------------------------------------------------------------------------------------
# 0. Choose soil type
for soil in ['sand', 'clay']:
    print(soil)
    if (soil == 'sand'):
        soildata = soildata_sand
        parameters = ParameterProvider(cropdata=cropdata, soildata=soildata, sitedata=sitedata)
    else:
        soildata = soildata_clay
        parameters = ParameterProvider(cropdata=cropdata, soildata=soildata, sitedata=sitedata)

# ----------------------------------------------------------------------------------------------------------------------
# 1. Loop per weather station
    weather_stations = ['DE BILT', 'EELDE', 'VLISSINGEN']
    for station in weather_stations:
        print(station)
        # 2. Select weather station
        weatherdata = ExcelWeatherDataProvider(os.path.join(input_dir, "weather", f"WOFOST_Weather_KNMI_{station}.xlsx"))
        # 3. Get co2 data
        mauna_loa = pd.read_csv(os.path.join(input_dir, "co2_annmean_mlo.csv"))
        mauna_loa = mauna_loa[mauna_loa.year > 1971]
        mauna_loa = mauna_loa[mauna_loa.year < 2017]
        CO2levels = list(mauna_loa['mean'])
        years = list(mauna_loa['year'])
        Output = pd.DataFrame()
        Summary = pd.DataFrame()
        for CO2, yr in zip(CO2levels, years):
            parameters.set_override("CO2", CO2)
            # 4. Load agromanager
            agrod = yaml.safe_load(open(os.path.join(input_dir, "agro-julius", f"Agrom_WW_Monocrop_{yr}.txt")))
            agrod = agrod['AgroManagement']
            # 5. Choose model configuration
            if (nlim == True and wlim == True):
                wofost = WOFOST8_ML_NWLP(parameters, weatherdata, agrod)
            elif (nlim == False and wlim == True):
                wofost = WOFOST8_ML_WLP(parameters, weatherdata, agrod)
            elif (nlim == True and wlim == False):
                wofost = WOFOST8_ML_NLP(parameters, weatherdata, agrod)
            else:
                wofost = WOFOST8_ML_PP(parameters, weatherdata, agrod)
            # 6. Run model
            wofost.run_till_terminate()
            # 7. Save model outputs
            Output_PP = pd.DataFrame(wofost.get_output())
            Summary_PP = pd.DataFrame(wofost.get_summary_output())
            # 8. Append years
            Output = Output.append(Output_PP)
            Summary = Summary.append(Summary_PP)
        # 9. Save daily output
        Output['dt_day'] = pd.to_datetime(Output.day)
        weather_df = pd.DataFrame(weatherdata.export())
        weather_df['dt_day'] = pd.to_datetime(weather_df.DAY)
        df_wofost_meteo = pd.merge(Output, weather_df, on='dt_day')
        path = os.path.join(output_dir, f"julius-2009/yw-long-term-daily-{station}-co2-{soil}.xlsx")
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_wofost_meteo.to_excel(writer)
        writer.save()
        writer.close()
        # 10. Save summary output
        Summary['DOM'] = pd.to_datetime(Summary['DOM'])
        Summary['year'] = Summary['DOM'].dt.year
        Summary = pd.merge(Summary, mauna_loa, on='year')
        path = os.path.join(output_dir, f"julius-2009/yw-long-term-summary-{station}-co2-{soil}.xlsx")
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        Summary.to_excel(writer)
        writer.save()
        writer.close()
